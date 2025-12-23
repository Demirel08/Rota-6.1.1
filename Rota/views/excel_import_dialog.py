"""
EFES ROTA X - Excel Sipariş Aktarma Dialogu
Excel dosyasından toplu sipariş yükleme
"""

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QTableWidget, QTableWidgetItem,
    QFrame, QMessageBox, QFileDialog, QProgressBar,
    QHeaderView, QCheckBox, QGroupBox, QWidget
)
from PySide6.QtCore import Qt, QDate, QThread, Signal
from PySide6.QtGui import QFont, QColor
from datetime import datetime
import openpyxl
from openpyxl import Workbook

try:
    from utils.timezone_helper import now_turkey, get_current_date_turkey
except ImportError:
    # Fallback: timezone_helper bulunamazsa normal datetime kullan
    now_turkey = lambda: datetime.now()
    get_current_date_turkey = lambda: datetime.now().date()

try:
    from core.db_manager import db
except ImportError:
    db = None


# =============================================================================
# TEMA
# =============================================================================
class Colors:
    BG = "#FFFFFF"
    HEADER_BG = "#F3F3F3"
    BORDER = "#D4D4D4"
    GRID = "#E0E0E0"
    TEXT = "#1A1A1A"
    TEXT_SECONDARY = "#666666"
    TEXT_MUTED = "#999999"
    ACCENT = "#217346"

    CRITICAL = "#C00000"
    WARNING = "#C65911"
    SUCCESS = "#107C41"
    INFO = "#0066CC"


# =============================================================================
# IMPORT WORKER THREAD (Performans için arka plan işlemi)
# =============================================================================
class ImportWorkerThread(QThread):
    """Sipariş import işlemini arka planda yapar"""

    progress_updated = Signal(int, int)  # (current, total)
    import_completed = Signal(int, int, list)  # (success_count, error_count, error_messages)

    def __init__(self, orders_data):
        super().__init__()
        self.orders_data = orders_data

    def run(self):
        """Thread'de çalışacak kod - Thread-safe veritabanı işlemi"""
        import sqlite3
        import os
        import sys
        from datetime import datetime as _dt

        try:
            from utils.timezone_helper import now_turkey
        except ImportError:
            now_turkey = lambda: _dt.now()

        try:
            from core.refresh_manager import refresh_manager
            from core.cache_manager import query_cache, order_cache, station_cache
        except ImportError:
            # Fallback
            class DummyCache:
                def invalidate_table(self, table): pass
                def clear(self): pass
            class DummyRefresh:
                def mark_dirty(self, key): pass
            refresh_manager = DummyRefresh()
            query_cache = DummyCache()
            order_cache = DummyCache()
            station_cache = DummyCache()

        success_count = 0
        error_count = 0
        error_messages = []

        try:
            # Thread içinde kendi veritabanı bağlantısını oluştur
            if getattr(sys, 'frozen', False):
                app_data = os.path.join(os.environ['LOCALAPPDATA'], 'REFLEKS360ROTA')
            else:
                app_data = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

            db_path = os.path.join(app_data, "efes_factory.db")

            # Thread-safe connection oluştur
            conn = sqlite3.connect(db_path, check_same_thread=False)
            conn.row_factory = sqlite3.Row

            created_time = now_turkey().strftime('%Y-%m-%d %H:%M:%S')

            # Transaction başlat (performans için)
            conn.execute("BEGIN TRANSACTION")

            for idx, data in enumerate(self.orders_data):
                try:
                    total_m2 = data.get('total_m2') or 0

                    # Sipariş ekle
                    conn.execute("""
                        INSERT INTO orders (order_code, customer_name, product_type, thickness, quantity,
                                           delivery_date, priority, status, route, declared_total_m2, width, height,
                                           sale_price, total_price, notes, project_id, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 'Beklemede', ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (data['code'], data['customer'], data['product'], data['thickness'], data['quantity'],
                          data['date'], data['priority'], data.get('route', ''), total_m2,
                          data.get('width', 0), data.get('height', 0), 0, 0,
                          data.get('notes', ''), data.get('project_id'), created_time))

                    # Stok düş
                    p_name = f"{data['thickness']}mm {data['product']}"
                    conn.execute("UPDATE stocks SET quantity_m2 = quantity_m2 - ? WHERE product_name = ?",
                               (total_m2, p_name))

                    success_count += 1

                except Exception as e:
                    error_count += 1
                    error_messages.append(f"{data.get('code', 'Bilinmeyen')}: {str(e)}")

                # Progress güncelle
                self.update_progress(idx + 1, len(self.orders_data))

            # Transaction'ı commit et
            conn.commit()
            conn.close()

            # Cache'i temizle (ana thread'de yapılmalı ama signal ile bildir)
            if success_count > 0:
                refresh_manager.mark_dirty('orders')
                refresh_manager.mark_dirty('stocks')
                query_cache.invalidate_table('orders')
                query_cache.invalidate_table('stocks')
                order_cache.clear()
                station_cache.clear()

            # Sonucu bildir
            self.import_completed.emit(success_count, error_count, error_messages)

        except Exception as e:
            # Hata durumunda
            import traceback
            error_detail = f"{str(e)}\n{traceback.format_exc()}"
            self.import_completed.emit(0, len(self.orders_data), [error_detail])

    def update_progress(self, current, total):
        """Progress callback - UI'ye bildirim gönder"""
        self.progress_updated.emit(current, total)


# =============================================================================
# EXCEL IMPORT DIALOG
# =============================================================================
class ExcelImportDialog(QDialog):
    """Excel'den Sipariş Aktarma Dialogu"""

    # Beklenen Excel sütun başlıkları (Türkçe)
    EXPECTED_COLUMNS = [
        "Sipariş Kodu",
        "Müşteri",
        "Ürün Tipi",
        "Kalınlık (mm)",
        "Adet",
        "Toplam m²",
        "Öncelik",
        "Teslim Tarihi",
        "Rota",
        "Not"
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Excel'den Sipariş Aktar")
        self.setMinimumSize(1000, 650)
        self.resize(1100, 700)

        self.excel_file_path = None
        self.orders_data = []

        self.setup_ui()

    def setup_ui(self):
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {Colors.BG};
            }}
            QLabel {{
                color: {Colors.TEXT};
                font-size: 11px;
            }}
            QTableWidget {{
                background-color: {Colors.BG};
                border: 1px solid {Colors.BORDER};
                gridline-color: {Colors.GRID};
                font-size: 10px;
            }}
            QTableWidget::item {{
                padding: 4px;
            }}
            QHeaderView::section {{
                background-color: {Colors.HEADER_BG};
                border: 1px solid {Colors.BORDER};
                padding: 6px;
                font-weight: bold;
                font-size: 10px;
            }}
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Header
        header = QFrame()
        header.setFixedHeight(50)
        header.setStyleSheet(f"""
            QFrame {{
                background-color: {Colors.HEADER_BG};
                border-bottom: 1px solid {Colors.BORDER};
            }}
        """)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(20, 0, 20, 0)

        title = QLabel("Excel'den Sipariş Aktar")
        title.setStyleSheet(f"font-size: 14px; font-weight: bold; color: {Colors.TEXT};")
        header_layout.addWidget(title)
        header_layout.addStretch()

        # Şablon indir butonu
        btn_download_template = QPushButton("📥 Örnek Şablon İndir")
        btn_download_template.setFixedSize(160, 32)
        btn_download_template.setCursor(Qt.PointingHandCursor)
        btn_download_template.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.INFO};
                border: none;
                border-radius: 4px;
                color: white;
                font-size: 10px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #0052A3;
            }}
        """)
        btn_download_template.clicked.connect(self.download_template)
        header_layout.addWidget(btn_download_template)

        layout.addWidget(header)

        # İçerik alanı
        content = QFrame()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(20, 20, 20, 20)
        content_layout.setSpacing(16)

        # Dosya seçimi bölümü
        file_group = QGroupBox("1. Excel Dosyası Seç")
        file_group.setStyleSheet(f"""
            QGroupBox {{
                font-size: 11px;
                font-weight: bold;
                color: {Colors.TEXT};
                border: 1px solid {Colors.BORDER};
                border-radius: 4px;
                margin-top: 12px;
                padding-top: 12px;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }}
        """)
        file_layout = QHBoxLayout(file_group)

        self.lbl_file_path = QLabel("Henüz dosya seçilmedi")
        self.lbl_file_path.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; font-weight: normal;")
        file_layout.addWidget(self.lbl_file_path, 1)

        btn_select_file = QPushButton("📂 Excel Dosyası Seç")
        btn_select_file.setFixedSize(150, 36)
        btn_select_file.setCursor(Qt.PointingHandCursor)
        btn_select_file.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.ACCENT};
                border: none;
                border-radius: 4px;
                color: white;
                font-size: 11px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #1D6640;
            }}
        """)
        btn_select_file.clicked.connect(self.select_excel_file)
        file_layout.addWidget(btn_select_file)

        content_layout.addWidget(file_group)

        # Önizleme tablosu
        preview_group = QGroupBox("2. Siparişleri Önizle ve Kontrol Et")
        preview_group.setStyleSheet(f"""
            QGroupBox {{
                font-size: 11px;
                font-weight: bold;
                color: {Colors.TEXT};
                border: 1px solid {Colors.BORDER};
                border-radius: 4px;
                margin-top: 12px;
                padding-top: 12px;
            }}
        """)
        preview_layout = QVBoxLayout(preview_group)

        self.table = QTableWidget()
        self.table.setColumnCount(11)
        self.table.setHorizontalHeaderLabels([
            "✓", "Sipariş Kodu", "Müşteri", "Ürün", "Kalınlık",
            "Adet", "m²", "Öncelik", "Teslim Tarihi", "Rota", "Not"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.setAlternatingRowColors(True)
        preview_layout.addWidget(self.table)

        # Seçim butonları
        selection_layout = QHBoxLayout()

        btn_select_all = QPushButton("Tümünü Seç")
        btn_select_all.setFixedHeight(28)
        btn_select_all.clicked.connect(self.select_all_rows)
        btn_select_all.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.BG};
                border: 1px solid {Colors.BORDER};
                border-radius: 3px;
                padding: 4px 12px;
                font-size: 10px;
            }}
            QPushButton:hover {{
                background-color: {Colors.HEADER_BG};
            }}
        """)
        selection_layout.addWidget(btn_select_all)

        btn_deselect_all = QPushButton("Tümünü Kaldır")
        btn_deselect_all.setFixedHeight(28)
        btn_deselect_all.clicked.connect(self.deselect_all_rows)
        btn_deselect_all.setStyleSheet(btn_select_all.styleSheet())
        selection_layout.addWidget(btn_deselect_all)

        selection_layout.addStretch()

        self.lbl_selected_count = QLabel("Seçili: 0 sipariş")
        self.lbl_selected_count.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; font-weight: normal;")
        selection_layout.addWidget(self.lbl_selected_count)

        preview_layout.addLayout(selection_layout)

        content_layout.addWidget(preview_group)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet(f"""
            QProgressBar {{
                border: 1px solid {Colors.BORDER};
                border-radius: 3px;
                text-align: center;
                font-size: 10px;
            }}
            QProgressBar::chunk {{
                background-color: {Colors.ACCENT};
            }}
        """)
        content_layout.addWidget(self.progress_bar)

        layout.addWidget(content)

        # Alt butonlar
        footer = QFrame()
        footer.setFixedHeight(60)
        footer.setStyleSheet(f"""
            QFrame {{
                background-color: {Colors.HEADER_BG};
                border-top: 1px solid {Colors.BORDER};
            }}
        """)
        footer_layout = QHBoxLayout(footer)
        footer_layout.setContentsMargins(20, 12, 20, 12)
        footer_layout.setSpacing(12)

        btn_cancel = QPushButton("İptal")
        btn_cancel.setFixedSize(100, 36)
        btn_cancel.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.BG};
                border: 1px solid {Colors.BORDER};
                border-radius: 4px;
                color: {Colors.TEXT_SECONDARY};
                font-size: 11px;
            }}
            QPushButton:hover {{
                background-color: {Colors.HEADER_BG};
            }}
        """)
        btn_cancel.clicked.connect(self.reject)
        footer_layout.addWidget(btn_cancel)

        footer_layout.addStretch()

        self.btn_import = QPushButton("Seçili Siparişleri Aktar")
        self.btn_import.setFixedHeight(36)
        self.btn_import.setMinimumWidth(180)
        self.btn_import.setCursor(Qt.PointingHandCursor)
        self.btn_import.setEnabled(False)
        self.btn_import.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.ACCENT};
                border: none;
                border-radius: 4px;
                padding: 0 24px;
                color: white;
                font-size: 12px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #1D6640;
            }}
            QPushButton:disabled {{
                background-color: {Colors.GRID};
                color: {Colors.TEXT_MUTED};
            }}
        """)
        self.btn_import.clicked.connect(self.import_orders)
        footer_layout.addWidget(self.btn_import)

        layout.addWidget(footer)

    def download_template(self):
        """Örnek Excel şablonunu indir"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Şablonu Kaydet",
            "siparis_sablonu.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Siparişler"

            # Başlıkları yaz
            headers = self.EXPECTED_COLUMNS
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="217346", end_color="217346", fill_type="solid")
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)

            # Örnek veri satırları ekle (bazı alanlar boş - test için)
            example_data = [
                ["SIP-001", "ABC Cam Ltd.", "Düz Cam", 6, 10, 5.5, "Acil", "31.12.2025", "KESİM,TEMPERLEME", "Tüm alanlar dolu"],
                ["SIP-002", "XYZ İnşaat", "Lamine Cam", 8, 5, "", "Normal", "", "", "Bazı alanlar boş - sorun yok"],
                ["SIP-003", "Demo Müşteri", "", "", 15, 3.2, "", "15.01.2026", "KESİM", "Ürün tipi ve kalınlık boş"],
            ]

            for row_idx, row_data in enumerate(example_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Sütun genişliklerini ayarla
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['K'].width = 35
            ws.column_dimensions['M'].width = 20

            # Açıklama sayfası ekle
            ws_info = wb.create_sheet("Kullanım Kılavuzu")
            info_text = [
                ["EFES ROTA X - SİPARİŞ AKTARMA ŞABLONU", ""],
                ["", ""],
                ["KULLANIM TALİMATLARI:", ""],
                ["1. 'Siparişler' sekmesindeki örnek satırları silin", ""],
                ["2. Kendi siparişlerinizi girin", ""],
                ["3. Tüm zorunlu alanları doldurun", ""],
                ["", ""],
                ["ALAN AÇIKLAMALARI:", ""],
                ["Sipariş Kodu:", "Benzersiz sipariş numarası (örn: I2025-001)"],
                ["Müşteri:", "Müşteri firma adı"],
                ["Ürün Tipi:", "Düz Cam, Temperli, Lamine, Satina, Renkli, Ayna (boş bırakılabilir)"],
                ["Kalınlık (mm):", "Cam kalınlığı (2-19 mm arası, boş bırakılabilir)"],
                ["Adet:", "Sipariş adedi (zorunlu)"],
                ["Toplam m²:", "Toplam metrekare (boş bırakılabilir)"],
                ["Öncelik:", "Normal, Acil, Çok Acil, Kritik (boş bırakılabilir)"],
                ["Teslim Tarihi:", "GG.AA.YYYY formatında (örn: 15.01.2025, boş bırakılabilir)"],
                ["Rota:", "İstasyonlar virgülle ayrılmış (örn: INTERMAC,DOUBLEDGER,SEVKIYAT, boş bırakılabilir)"],
                ["Not:", "Ek notlar (boş bırakılabilir)"],
                ["", ""],
                ["ROTADA KULLANILABİLECEK İSTASYONLAR:", ""],
                ["Kesim:", "INTERMAC, LIVA KESIM, LAMINE KESIM"],
                ["İşleme:", "CNC RODAJ, DOUBLEDGER, ZIMPARA, DELIK, OYGU"],
                ["Tesir:", "TESIR A1, TESIR B1, TESIR B1-1, TESIR B1-2"],
                ["Fırın:", "TEMPER A1, TEMPER B1, TEMPER BOMBE"],
                ["Montaj:", "LAMINE A1, ISICAM B1, KUMLAMA"],
                ["Sevkiyat:", "SEVKIYAT (otomatik eklenir)"],
            ]

            for row_idx, row_data in enumerate(info_text, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:
                        cell.font = openpyxl.styles.Font(bold=True, size=14, color="217346")
                    elif col_idx == 1 and value and ":" in value:
                        cell.font = openpyxl.styles.Font(bold=True)

            ws_info.column_dimensions['A'].width = 25
            ws_info.column_dimensions['B'].width = 50

            wb.save(file_path)
            QMessageBox.information(self, "Başarılı", f"Şablon dosyası oluşturuldu:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Şablon oluşturulamadı:\n{str(e)}")

    def select_excel_file(self):
        """Excel dosyası seç ve oku"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Excel Dosyası Seç",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if not file_path:
            return

        self.excel_file_path = file_path
        self.lbl_file_path.setText(file_path)

        # Excel dosyasını oku ve tabloya yükle
        self.load_excel_data()

    def _get_cell_value(self, cell, column_name):
        """Hücre değerini tipine göre oku ve dönüştür"""
        from datetime import datetime
        import openpyxl.cell.cell

        if cell.value is None or str(cell.value).strip() == "":
            return None

        # Tarih sütunu
        if column_name == "Teslim Tarihi":
            if isinstance(cell.value, datetime):
                return cell.value.strftime("%d.%m.%Y")
            else:
                return str(cell.value).strip()

        # Sayı sütunları
        elif column_name in ["Adet", "Kalınlık (mm)"]:
            try:
                # Eğer Excel tarihe çevirdiyse, sayıya geri dön
                if isinstance(cell.value, datetime):
                    # Excel serial date'i sayıya çevir
                    return int(cell.value.day) if column_name == "Adet" else int(cell.value.month)
                return int(float(cell.value))
            except:
                return None

        # Ondalık sayı sütunları
        elif column_name == "Toplam m²":
            try:
                if isinstance(cell.value, datetime):
                    return 0.0
                return float(cell.value)
            except:
                return None

        # Metin sütunları
        else:
            return str(cell.value).strip()

    def load_excel_data(self):
        """Excel dosyasından verileri oku ve tabloya yükle"""
        if not self.excel_file_path:
            return

        try:
            wb = openpyxl.load_workbook(self.excel_file_path, data_only=True)

            # "Siparişler" sayfasını ara, yoksa ilk sayfayı al
            if "Siparişler" in wb.sheetnames:
                ws = wb["Siparişler"]
            else:
                ws = wb.active

            # Başlıkları oku ve eşleştir
            excel_headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

            # Kolon eşleştirme map'i oluştur
            column_map = {}
            for expected_col in self.EXPECTED_COLUMNS:
                # Tam eşleşme ara
                if expected_col in excel_headers:
                    column_map[expected_col] = excel_headers.index(expected_col)
                else:
                    # Büyük/küçük harf duyarsız eşleşme
                    expected_lower = expected_col.lower()
                    for idx, header in enumerate(excel_headers):
                        if header.lower() == expected_lower:
                            column_map[expected_col] = idx
                            break

            # Verileri oku
            self.orders_data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                # Boş satırları atla
                if not any(cell.value for cell in row):
                    continue

                order_dict = {}

                # Her beklenen sütun için veriyi al
                for expected_col in self.EXPECTED_COLUMNS:
                    if expected_col in column_map:
                        col_idx = column_map[expected_col]
                        if col_idx < len(row):
                            cell = row[col_idx]
                            # Hücre tipine göre değer al
                            value = self._get_cell_value(cell, expected_col)
                            order_dict[expected_col] = value
                        else:
                            order_dict[expected_col] = None
                    else:
                        order_dict[expected_col] = None

                order_dict['_row_number'] = row_idx
                order_dict['_valid'] = True
                order_dict['_errors'] = []

                # Basit validasyon
                if not order_dict.get("Sipariş Kodu"):
                    order_dict['_valid'] = False
                    order_dict['_errors'].append("Sipariş kodu eksik")

                if not order_dict.get("Müşteri"):
                    order_dict['_valid'] = False
                    order_dict['_errors'].append("Müşteri adı eksik")

                try:
                    adet = int(order_dict.get("Adet", 0))
                    if adet <= 0:
                        order_dict['_valid'] = False
                        order_dict['_errors'].append("Adet 0'dan büyük olmalı")
                except:
                    order_dict['_valid'] = False
                    order_dict['_errors'].append("Adet sayı olmalı")

                self.orders_data.append(order_dict)

            # Tabloya yükle
            self.populate_table()

            if self.orders_data:
                self.btn_import.setEnabled(True)
                QMessageBox.information(
                    self,
                    "Başarılı",
                    f"{len(self.orders_data)} sipariş okundu.\n"
                    f"Geçerli: {sum(1 for o in self.orders_data if o['_valid'])}\n"
                    f"Hatalı: {sum(1 for o in self.orders_data if not o['_valid'])}"
                )
            else:
                QMessageBox.warning(self, "Uyarı", "Excel dosyasında sipariş bulunamadı!")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel dosyası okunamadı:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def populate_table(self):
        """Verileri tabloya yükle"""
        self.table.setRowCount(len(self.orders_data))

        for row_idx, order in enumerate(self.orders_data):
            # Checkbox
            chk = QCheckBox()
            chk.setChecked(order['_valid'])  # Geçerli olanları otomatik seç
            chk.stateChanged.connect(self.update_selected_count)
            chk_widget = QWidget()
            chk_layout = QHBoxLayout(chk_widget)
            chk_layout.addWidget(chk)
            chk_layout.setAlignment(Qt.AlignCenter)
            chk_layout.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(row_idx, 0, chk_widget)

            # Veriler (sadeleştirilmiş sütunlar)
            columns_data = [
                order.get("Sipariş Kodu", ""),
                order.get("Müşteri", ""),
                order.get("Ürün Tipi", ""),
                str(order.get("Kalınlık (mm)", "")) if order.get("Kalınlık (mm)") else "",
                str(order.get("Adet", "")) if order.get("Adet") else "",
                str(order.get("Toplam m²", "")) if order.get("Toplam m²") else "",
                order.get("Öncelik", ""),
                str(order.get("Teslim Tarihi", "")) if order.get("Teslim Tarihi") else "",
                order.get("Rota", ""),
                order.get("Not", ""),
            ]

            for col_idx, data in enumerate(columns_data, start=1):
                item = QTableWidgetItem(str(data) if data else "")

                # Hatalı satırları kırmızı yap
                if not order['_valid']:
                    item.setBackground(QColor("#FFE6E6"))
                    item.setForeground(QColor(Colors.CRITICAL))
                    if col_idx == 1:  # Sipariş kodu sütununa hata mesajı ekle
                        item.setToolTip("\n".join(order['_errors']))

                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row_idx, col_idx, item)

        self.update_selected_count()

    def select_all_rows(self):
        """Tüm satırları seç"""
        for row in range(self.table.rowCount()):
            chk_widget = self.table.cellWidget(row, 0)
            if chk_widget:
                chk = chk_widget.findChild(QCheckBox)
                if chk:
                    chk.setChecked(True)

    def deselect_all_rows(self):
        """Tüm seçimleri kaldır"""
        for row in range(self.table.rowCount()):
            chk_widget = self.table.cellWidget(row, 0)
            if chk_widget:
                chk = chk_widget.findChild(QCheckBox)
                if chk:
                    chk.setChecked(False)

    def update_selected_count(self):
        """Seçili sipariş sayısını güncelle"""
        count = 0
        for row in range(self.table.rowCount()):
            chk_widget = self.table.cellWidget(row, 0)
            if chk_widget:
                chk = chk_widget.findChild(QCheckBox)
                if chk and chk.isChecked():
                    count += 1

        self.lbl_selected_count.setText(f"Seçili: {count} sipariş")

    def import_orders(self):
        """Seçili siparişleri veritabanına aktar (Thread kullanarak)"""
        if not db:
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantısı yok!")
            return

        # Seçili siparişleri topla
        selected_orders = []
        for row in range(self.table.rowCount()):
            chk_widget = self.table.cellWidget(row, 0)
            if chk_widget:
                chk = chk_widget.findChild(QCheckBox)
                if chk and chk.isChecked():
                    selected_orders.append(self.orders_data[row])

        if not selected_orders:
            QMessageBox.warning(self, "Uyarı", "Lütfen en az bir sipariş seçin!")
            return

        # Onay al
        reply = QMessageBox.question(
            self,
            "Onay",
            f"{len(selected_orders)} sipariş sisteme aktarılacak.\n\nDevam edilsin mi?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        # Boş değerleri güvenli şekilde handle et
        def safe_int(value, default=0):
            if value is None or value == "" or str(value).strip() == "":
                return default
            try:
                return int(float(value))
            except:
                return default

        def safe_float(value, default=0.0):
            if value is None or value == "" or str(value).strip() == "":
                return default
            try:
                return float(value)
            except:
                return default

        def safe_str(value, default=""):
            if value is None or str(value).strip() == "":
                return default
            return str(value).strip()

        # Siparişleri hazırla
        prepared_orders = []
        for order in selected_orders:
            # Tarihi formatla
            delivery_date = order.get("Teslim Tarihi", "")
            if delivery_date:
                try:
                    if isinstance(delivery_date, datetime):
                        delivery_date = delivery_date.strftime("%Y-%m-%d")
                    elif isinstance(delivery_date, str):
                        parts = delivery_date.split('.')
                        if len(parts) == 3:
                            delivery_date = f"{parts[2]}-{parts[1]}-{parts[0]}"
                except:
                    delivery_date = QDate.currentDate().addDays(7).toString("yyyy-MM-dd")
            else:
                delivery_date = QDate.currentDate().addDays(7).toString("yyyy-MM-dd")

            # Rotayı düzenle
            route = order.get("Rota", "").strip()
            if route and "SEVKIYAT" not in route:
                route = route + ",SEVKIYAT"
            elif not route:
                route = "SEVKIYAT"

            # Sipariş verisini hazırla
            order_data = {
                "code": safe_str(order.get("Sipariş Kodu"), ""),
                "customer": safe_str(order.get("Müşteri"), ""),
                "product": safe_str(order.get("Ürün Tipi"), "Düz Cam"),
                "thickness": safe_int(order.get("Kalınlık (mm)"), 6),
                "width": safe_float(order.get("Genişlik (mm)"), 0),
                "height": safe_float(order.get("Yükseklik (mm)"), 0),
                "quantity": safe_int(order.get("Adet"), 1),
                "total_m2": safe_float(order.get("Toplam m²"), 0),
                "priority": safe_str(order.get("Öncelik"), "Normal"),
                "date": delivery_date,
                "route": route,
                "sale_price": safe_float(order.get("Satış Fiyatı"), 0),
                "notes": safe_str(order.get("Not"), ""),
                "project_id": None
            }
            prepared_orders.append(order_data)

        # Progress bar göster
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(len(prepared_orders))
        self.progress_bar.setValue(0)

        # Butonları devre dışı bırak
        self.btn_import.setEnabled(False)

        # Worker thread'i başlat
        self.worker = ImportWorkerThread(prepared_orders)
        self.worker.progress_updated.connect(self.on_progress_updated)
        self.worker.import_completed.connect(self.on_import_completed)
        self.worker.start()

    def on_progress_updated(self, current, total):
        """Progress bar güncelleme"""
        self.progress_bar.setValue(current)

    def on_import_completed(self, success_count, error_count, error_messages):
        """Import tamamlandığında çağrılır"""
        # Progress bar gizle
        self.progress_bar.setVisible(False)
        self.btn_import.setEnabled(True)

        # Sonucu göster
        result_msg = f"Aktarım tamamlandı!\n\n"
        result_msg += f"Başarılı: {success_count}\n"
        result_msg += f"Başarısız: {error_count}"

        if error_messages:
            result_msg += "\n\nHatalar:\n" + "\n".join(error_messages[:5])
            if len(error_messages) > 5:
                result_msg += f"\n... ve {len(error_messages) - 5} hata daha"

        if error_count > 0:
            QMessageBox.warning(self, "Tamamlandı", result_msg)
        else:
            QMessageBox.information(self, "Başarılı", result_msg)

        # Dialog'u kapat
        if success_count > 0:
            self.accept()


if __name__ == "__main__":
    from PySide6.QtWidgets import QApplication
    import sys

    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 9))

    dialog = ExcelImportDialog()
    dialog.show()

    sys.exit(app.exec())
